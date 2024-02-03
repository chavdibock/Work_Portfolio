import openpyxl as xl
from openpyxl.styles import Alignment
import re
import os
from datetime import datetime

path = "C:/Users/User/PycharmProjects/SDP_NUMBERING_CONVENTION/configs"


wb = xl.Workbook()
ws = wb.active

ws.title = "sdp_numbering"

ws.merge_cells("A1:E2")
ws['A1'] = str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15

os.chdir(path)
for file in os.listdir():
    # Check whether file is in text format or not
    list_config_device = []
    sdp_number = []
    if file.endswith(".setconf"):
        file_path = f"{path}/{file}"
        new_name = file.split(".")
        last_taken_row = str(ws.max_row + 1)
        ws.merge_cells("A" + last_taken_row + ":" + "E" + last_taken_row)
        ws["A" + last_taken_row] = new_name[0]
        ws['A' + last_taken_row].alignment = Alignment(horizontal='center', vertical='center')
        with open(file_path, 'r') as device_config:
            for i in device_config:
                list_config_device.append(i)

            sdp_reg = r'/configure service sdp (.*) create'

            for x in list_config_device:
                math_sdp = re.match(sdp_reg, x)
                if math_sdp and math_sdp.group(1) not in sdp_number:
                    serv_id = math_sdp.group(1).split(" ")
                    if len(serv_id) == 2:
                        sdp_number.append(serv_id[0])
                        print(serv_id[0])

            sdp_far_end_list = []
            sdp_descpr_list = []
            real_service_id = []

            for i in sdp_number:

                sdp_far_end = r'/configure service sdp {0} (.*) far-end (.*)'.format(re.escape(i))
                sdp_desc = r'/configure service sdp {0} (.*) description (.*)'.format(re.escape(i))

                size_dscp_list = len(sdp_descpr_list)
                for n in list_config_device:
                    match_far_end = re.match(sdp_far_end, n)
                    match_desc = re.match(sdp_desc, n)

                    if match_far_end:
                        sdp_far_end_list.append(match_far_end.group(2))
                        ip_octets = match_far_end.group(2).split(".")
                        real_service_id.append(str(int(ip_octets[2]) * 1000 + int(ip_octets[3])))
                        print(match_far_end.group(2))
                    elif match_desc:
                        sdp_descpr_list.append(match_desc.group(2))
                        print(match_desc.group(2))

                new_size_dscp_list = len(sdp_descpr_list)

                if size_dscp_list == new_size_dscp_list:
                    sdp_descpr_list.append("no description")

            print("service len  " + str(len(sdp_number)))
            print("descpr len  " + str(len(sdp_descpr_list)))
            print("far_end " + str(len(sdp_far_end_list)))
            print("real_service " + str(len(real_service_id)))

            used_sdp = []
            not_used_sdp = []
            invalid_sdp_numb = []
            check = []

            for x in range(len(sdp_number)):
                sdp_vpls = r'/configure service vpls (.*)-sdp {0}(.*) create'.format(sdp_number[x])
                sdp_epipe = r'/configure service epipe (.*)-sdp {0}(.*) create'.format(sdp_number[x])
                convention = "VALID"
                if sdp_number[x] != real_service_id[x]:
                    convention = "INVALID"
                    invalid_sdp_numb.append("invalid for " + sdp_number[x] + " and " + sdp_far_end_list[x])

                size_used_sdp = len(used_sdp)
                for i in list_config_device:
                    match_isUsed_vpls = re.match(sdp_vpls, i)
                    match_isUsed_epipe = re.match(sdp_epipe, i)

                    if match_isUsed_epipe and sdp_number[x] not in check:
                        used_sdp.append(sdp_number[x] + " " + sdp_far_end_list[x] + " " + sdp_descpr_list[x])
                        check.append(sdp_number[x])
                        ws.append([sdp_number[x], sdp_descpr_list[x], sdp_far_end_list[x], convention, 'USED'])

                    elif match_isUsed_vpls and sdp_number[x] not in check:
                        used_sdp.append(sdp_number[x] + " " + sdp_far_end_list[x] + " " + sdp_descpr_list[x])
                        check.append(sdp_number[x])
                        ws.append([sdp_number[x], sdp_descpr_list[x], sdp_far_end_list[x], convention, 'USED'])

                new_size_used_sdp = len(used_sdp)

                if size_used_sdp == new_size_used_sdp and sdp_number[x] not in check:
                    not_used_sdp.append(sdp_number[x] + " " + sdp_far_end_list[x])
                    check.append(sdp_number[x])
                    ws.append([sdp_number[x], sdp_descpr_list[x], sdp_far_end_list[x], convention, 'UNUSED'])

            wb.save("table.xlsx")
