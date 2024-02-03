from datetime import datetime
import openpyxl as xl

wb = xl.Workbook()
ws = wb.active

ws.title = "sdp_numbering"
ws.merge_cells("A1:E2")
ws['A1'] = str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
wb.save("table.xlsx")
