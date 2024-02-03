import openpyxl as xl
from openpyxl.styles import Alignment
import re
import os
from datetime import datetime

wb = xl.Workbook()
ws = wb.active

ws.title = "sdp_numbering"

ws.merge_cells("A1:G1")
ws['A1'] = str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.column_dimensions['A'].width = 15  # sevice - epipe/VPLS
ws.row_dimensions[1].height = 25  # adjusting the height of the DATE row
ws.column_dimensions['B'].width = 15  # id
ws.column_dimensions['C'].width = 35  # name
ws.column_dimensions['D'].width = 15  # port
ws.column_dimensions['E'].width = 35  # serv_descpr
ws.column_dimensions['F'].width = 35  # port_descpr
ws.column_dimensions['G'].width = 35  # staus
path = "/Configs_2023_22_08/"
os.chdir(path)
path_for_new_config = "C:/Users/User/PycharmProjects/Lack_of_descriptions/new_configs_vpls_epipe/"

VPLS = r'/configure\s+service\s+(\S+)\s+(.*)\s+name\s+"(.*)"\s+customer\s+(.*?)\s+sap\s+(.*)\s+create'

for file in os.listdir():
    vpls_srvID = []
    vpls_customer = []
    vpls_name = []
    vpls_sap_int = []
    type_of_servise = []
    sap_int_to_be_removed = []
    physical_port_name = []
    physical_port_db = []
    list_vpls_create_descpr = []
    list_device_config = []
    # Check whether file is in text format or not

    if file.endswith(".setconf"):
        file_path = f"{path}/{file}"
        new_name = file.split(".")
        with open(file_path, 'r') as device_config:
            result = open(path_for_new_config + new_name[0] + "_newconfig_VPLS_EPIPE.txt", "w+")
            print("##################### " + new_name[0] + " #####################")
            last_taken_row = str(ws.max_row + 1)
            ws.merge_cells("A" + last_taken_row + ":" + "G" + last_taken_row)
            ws["A" + last_taken_row] = new_name[0]
            ws['A' + last_taken_row].alignment = Alignment(horizontal='center', vertical='center')
            last_taken_row = str(ws.max_row + 1)
            ws['A' + last_taken_row] = "service"
            ws['A' + last_taken_row].alignment = Alignment(horizontal='center', vertical='center')
            ws['B' + last_taken_row] = "ID"
            ws['B' + last_taken_row].alignment = Alignment(horizontal='center', vertical='center')
            ws['C' + last_taken_row] = "name"
            ws['C' + last_taken_row].alignment = Alignment(horizontal='center', vertical='center')
            ws['D' + last_taken_row] = "port"
            ws['D' + last_taken_row].alignment = Alignment(horizontal='center', vertical='center')
            ws['E' + last_taken_row] = "serv_descpr"
            ws['E' + last_taken_row].alignment = Alignment(horizontal='center', vertical='center')
            ws['F' + last_taken_row] = "port_db_descpr"
            ws['F' + last_taken_row].alignment = Alignment(horizontal='center', vertical='center')
            ws['G' + last_taken_row] = "status"
            ws['G' + last_taken_row].alignment = Alignment(horizontal='center', vertical='center')
            for x in device_config:
                list_device_config.append(x)

            for i in list_device_config:
                match = re.match(VPLS, i)
                if match and (match.group(1) == "vpls" or match.group(1) == "epipe") and match.group(5) not in vpls_sap_int:
                    # print(match.group(1) + " " + match.group(2) + " " + match.group(3) + " " + match.group(4) + " " + match.group(5))
                    type_of_servise.append(match.group(1))
                    vpls_srvID.append(match.group(2))
                    vpls_name.append(match.group(3))
                    vpls_customer.append(match.group(4))
                    vpls_sap_int.append(match.group(5))
                    physical_port = match.group(5).split(":")
                    if physical_port[0].startswith("lag-"):
                        physical_port_name.append(physical_port[0].replace("-", " "))
                    else:
                        physical_port_name.append("port " + physical_port[0])

            for x in range(len(physical_port_name)):
                db_descpr = r'/configure {0} description ".*, db=(.*?),.*"'.format(re.escape(physical_port_name[x]))
                size = len(physical_port_db)
                for i in list_device_config:
                    match_desc = re.match(db_descpr, i)
                    if match_desc:
                        physical_port_db.append(match_desc.group(1))
                        # print(match_desc.group(1))
                new_size = len(physical_port_db)
                if size == new_size:
                    physical_port_db.append("")
                    print("no db on " + physical_port_name[x])

            for x in range(len(vpls_sap_int)):
                description_patter = r'/configure service {0} {1} name "{2}" customer {3} sap {4} description "(.*)"'.format(type_of_servise[x],
                                                                                                                             vpls_srvID[x],
                                                                                                                             re.escape(vpls_name[x]),
                                                                                                                             vpls_customer[x],
                                                                                                                             re.escape(
                                                                                                                                 vpls_sap_int[x]))

                vpls_create_descpr = r'/configure service {0} {1} name "{2}" customer {3} description "(.*)"'.format(type_of_servise[x],
                                                                                                                     vpls_srvID[x],
                                                                                                                     re.escape(vpls_name[x]),
                                                                                                                     vpls_customer[x], )

                list_create_dscpr_size = len(list_vpls_create_descpr)
                for i in list_device_config:
                    match_des = re.match(description_patter, i)
                    match_vpls_create_description = re.match(vpls_create_descpr, i)
                    if match_des:
                        print(match_des.group(1))
                        sap_int_to_be_removed.append(vpls_sap_int[x])
                    elif match_vpls_create_description:
                        list_vpls_create_descpr.append(match_vpls_create_description.group(1))
                        print(match_vpls_create_description.group(1))
                list_create_dscpr_new_size = len(list_vpls_create_descpr)
                if list_create_dscpr_size == list_create_dscpr_new_size:
                    list_vpls_create_descpr.append(vpls_name[x])
                    print("no despription on service " + type_of_servise[x] + " " + vpls_srvID[x] + " name " + vpls_name[x] + " customer " +
                          vpls_customer[x])
                    print('/configure service {0} {1} name {2} customer {3} sap {4} description {2}'.format(type_of_servise[x],
                                                                                                            vpls_srvID[x],
                                                                                                            vpls_name[x],
                                                                                                            vpls_customer[x],
                                                                                                            vpls_sap_int[x]))

            for n in range(len(vpls_sap_int)):
                if vpls_sap_int[n] not in sap_int_to_be_removed:
                    if physical_port_db[n] == "" or physical_port_db[n] == "unmanaged":
                        print('/configure service {0} {1} name {2} customer {3} sap {4} description "{5}"'.format(type_of_servise[n],
                                                                                                                  vpls_srvID[n],
                                                                                                                  vpls_name[n],
                                                                                                                  vpls_customer[n],
                                                                                                                  vpls_sap_int[n],
                                                                                                                  list_vpls_create_descpr[n]))

                        result.write('/configure service {0} {1} name "{2}" customer {3} sap {4} description "{5}"\n'.format(type_of_servise[n],
                                                                                                                             vpls_srvID[n],
                                                                                                                             vpls_name[n],
                                                                                                                             vpls_customer[n],
                                                                                                                             vpls_sap_int[n],
                                                                                                                             list_vpls_create_descpr[n]
                                                                                                                             ))
                    else:

                        desc = list_vpls_create_descpr[n] + " -> " + physical_port_db[n]

                        print('/configure service {0} {1} name "{2}" customer {3} sap {4} description {5}'.format(type_of_servise[n],
                                                                                                                  vpls_srvID[n],
                                                                                                                  vpls_name[n],
                                                                                                                  vpls_customer[n],
                                                                                                                  vpls_sap_int[n],
                                                                                                                  desc))

                        result.write('/configure service {0} {1} name "{2}" customer {3} sap {4} description "{5}"\n'.format(type_of_servise[n],
                                                                                                                             vpls_srvID[n],
                                                                                                                             vpls_name[n],
                                                                                                                             vpls_customer[n],
                                                                                                                             vpls_sap_int[n],
                                                                                                                             desc))

            result.close()

os.chdir("C:/Users/User/PycharmProjects/Lack_of_descriptions/")
##########
wb.save("VPLS_decpr.xlsx")
